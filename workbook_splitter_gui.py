# -*- coding: utf-8 -*-
from __future__ import annotations

import multiprocessing
import os
import re
import sys
import threading
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook

APP_NAME = "Excel工作簿拆分器"
APP_VERSION = "1.0.0"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


@dataclass(frozen=True)
class SheetInfo:
    name: str
    state: str


def safe_filename(name: str, fallback: str = "工作表") -> str:
    cleaned = INVALID_FILENAME_CHARS.sub("_", name).strip().rstrip(". ")
    return cleaned or fallback


def unique_output_path(folder: Path, stem: str, suffix: str) -> Path:
    candidate = folder / f"{stem}{suffix}"
    number = 2
    while candidate.exists():
        candidate = folder / f"{stem} ({number}){suffix}"
        number += 1
    return candidate


def list_sheets(workbook_path: str | Path) -> list[SheetInfo]:
    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"文件不存在：{path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx、.xlsm、.xltx、.xltm 文件。")

    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    wb = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    try:
        return [SheetInfo(ws.title, ws.sheet_state) for ws in wb.worksheets]
    finally:
        wb.close()


def split_workbook(
    workbook_path: str | Path,
    output_folder: str | Path,
    selected_sheets: Iterable[str],
    progress_callback: Callable[[int, int, str, str], None] | None = None,
) -> list[Path]:
    source = Path(workbook_path)
    output_dir = Path(output_folder)
    selected = list(dict.fromkeys(selected_sheets))

    if not source.is_file():
        raise FileNotFoundError(f"源文件不存在：{source}")
    if not selected:
        raise ValueError("请至少选择一个工作表。")
    if source.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx、.xlsm、.xltx、.xltm 文件。")

    output_dir.mkdir(parents=True, exist_ok=True)
    keep_vba = source.suffix.lower() in {".xlsm", ".xltm"}
    output_suffix = ".xlsm" if keep_vba else ".xlsx"
    source_stem = safe_filename(source.stem, "工作簿")
    outputs: list[Path] = []

    for index, sheet_name in enumerate(selected, 1):
        if progress_callback:
            progress_callback(index - 1, len(selected), sheet_name, "正在处理")

        wb = load_workbook(
            source,
            read_only=False,
            data_only=False,
            keep_vba=keep_vba,
            keep_links=True,
        )
        try:
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"未找到工作表：{sheet_name}")

            target = wb[sheet_name]
            target.sheet_state = "visible"

            for ws in list(wb.worksheets):
                if ws.title != sheet_name:
                    wb.remove(ws)

            output_stem = f"{source_stem}_{safe_filename(sheet_name)}"
            output_path = unique_output_path(output_dir, output_stem, output_suffix)
            wb.save(output_path)
            outputs.append(output_path)
        finally:
            wb.close()

        if progress_callback:
            progress_callback(index, len(selected), sheet_name, "已完成")

    return outputs


class ScrollableSheetList(ttk.Frame):
    def __init__(self, master: tk.Misc) -> None:
        super().__init__(master)
        self.canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.window = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.inner.bind("<Configure>", self._sync_scroll_region)
        self.canvas.bind("<Configure>", self._sync_width)
        self.canvas.bind_all("<MouseWheel>", self._mousewheel)

    def _sync_scroll_region(self, _event=None) -> None:
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _sync_width(self, event) -> None:
        self.canvas.itemconfigure(self.window, width=event.width)

    def _mousewheel(self, event) -> None:
        try:
            self.canvas.yview_scroll(int(-event.delta / 120), "units")
        except tk.TclError:
            pass


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_NAME} {APP_VERSION}")
        self.geometry("820x680")
        self.minsize(720, 590)
        self.option_add("*Font", ("Microsoft YaHei UI", 10))

        try:
            icon_path = Path(getattr(sys, "_MEIPASS", Path(__file__).parent)) / "app.ico"
            if icon_path.exists():
                self.iconbitmap(str(icon_path))
        except Exception:
            pass

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.include_hidden_var = tk.BooleanVar(False)
        self.status_var = tk.StringVar("请选择需要拆分的 Excel 工作簿。")
        self.sheet_infos: list[SheetInfo] = []
        self.sheet_vars: dict[str, tk.BooleanVar] = {}
        self.running = False

        self._create_styles()
        self._build_ui()

    def _create_styles(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Title.TLabel", font=("Microsoft YaHei UI", 21, "bold"))
        style.configure("Sub.TLabel", foreground="#666666")
        style.configure("Primary.TButton", font=("Microsoft YaHei UI", 10, "bold"))

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=24)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(1, weight=1)
        outer.rowconfigure(5, weight=1)

        ttk.Label(outer, text="拆分工作簿", style="Title.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w"
        )
        ttk.Label(
            outer,
            text="将选中的工作表分别保存为独立的 Excel 文件",
            style="Sub.TLabel",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 22))

        ttk.Label(outer, text="源工作簿：").grid(row=2, column=0, sticky="w", pady=7)
        self.input_entry = ttk.Entry(outer, textvariable=self.input_var)
        self.input_entry.grid(row=2, column=1, sticky="ew", padx=8, pady=7)
        self.input_button = ttk.Button(outer, text="浏览...", command=self.choose_input)
        self.input_button.grid(row=2, column=2, sticky="ew", pady=7)

        toolbar = ttk.Frame(outer)
        toolbar.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(14, 7))
        toolbar.columnconfigure(0, weight=1)
        ttk.Label(toolbar, text="请选择待拆分的工作表：").grid(row=0, column=0, sticky="w")
        self.hidden_check = ttk.Checkbutton(
            toolbar,
            text="包含隐藏工作表",
            variable=self.include_hidden_var,
            command=self.refresh_sheet_list,
        )
        self.hidden_check.grid(row=0, column=1, padx=8)
        self.select_button = ttk.Button(toolbar, text="全选", width=8, command=self.select_all)
        self.select_button.grid(row=0, column=2, padx=3)
        self.clear_button = ttk.Button(toolbar, text="清空", width=8, command=self.clear_all)
        self.clear_button.grid(row=0, column=3, padx=3)

        border = ttk.Frame(outer, relief="solid", borderwidth=1)
        border.grid(row=5, column=0, columnspan=3, sticky="nsew")
        border.columnconfigure(0, weight=1)
        border.rowconfigure(0, weight=1)
        self.sheet_list = ScrollableSheetList(border)
        self.sheet_list.grid(row=0, column=0, sticky="nsew")
        self.show_hint("选择工作簿后，将在此处显示工作表。")

        ttk.Label(outer, text="保存位置：").grid(row=6, column=0, sticky="w", pady=(18, 7))
        self.output_entry = ttk.Entry(outer, textvariable=self.output_var)
        self.output_entry.grid(row=6, column=1, sticky="ew", padx=8, pady=(18, 7))
        self.output_button = ttk.Button(outer, text="浏览...", command=self.choose_output)
        self.output_button.grid(row=6, column=2, sticky="ew", pady=(18, 7))

        self.progress = ttk.Progressbar(outer, maximum=100, mode="determinate")
        self.progress.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(12, 5))

        bottom = ttk.Frame(outer)
        bottom.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(4, 0))
        bottom.columnconfigure(0, weight=1)
        ttk.Label(bottom, textvariable=self.status_var, foreground="#555555").grid(
            row=0, column=0, sticky="w"
        )
        self.start_button = ttk.Button(
            bottom, text="开始拆分", style="Primary.TButton", command=self.start_split
        )
        self.start_button.grid(row=0, column=1, padx=(8, 8), ipadx=12, ipady=5)
        self.exit_button = ttk.Button(bottom, text="退出", command=self.destroy)
        self.exit_button.grid(row=0, column=2, ipadx=8, ipady=5)

    def show_hint(self, text: str) -> None:
        for child in self.sheet_list.inner.winfo_children():
            child.destroy()
        ttk.Label(self.sheet_list.inner, text=text, foreground="#888888").pack(
            anchor="center", pady=80
        )

    def choose_input(self) -> None:
        path = filedialog.askopenfilename(
            title="选择 Excel 工作簿",
            filetypes=[
                ("Excel 工作簿", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("所有文件", "*.*"),
            ],
        )
        if not path:
            return
        self.input_var.set(path)
        self.output_var.set(str(Path(path).parent / f"{Path(path).stem}_拆分结果"))
        self.load_sheets()

    def choose_output(self) -> None:
        path = filedialog.askdirectory(title="选择保存文件夹")
        if path:
            self.output_var.set(path)

    def load_sheets(self) -> None:
        try:
            self.sheet_infos = list_sheets(self.input_var.get().strip())
            self.refresh_sheet_list()
            self.status_var.set(f"已读取 {len(self.sheet_infos)} 个工作表。")
        except Exception as exc:
            self.sheet_infos = []
            self.show_hint("工作簿读取失败。")
            messagebox.showerror("读取失败", str(exc))

    def refresh_sheet_list(self) -> None:
        previous = {name: var.get() for name, var in self.sheet_vars.items()}
        for child in self.sheet_list.inner.winfo_children():
            child.destroy()
        self.sheet_vars.clear()

        infos = [
            info for info in self.sheet_infos
            if self.include_hidden_var.get() or info.state == "visible"
        ]
        if not infos:
            self.show_hint("没有可显示的工作表。")
            return

        for row, info in enumerate(infos):
            selected = previous.get(info.name, info.state == "visible")
            var = tk.BooleanVar(selected)
            self.sheet_vars[info.name] = var
            suffix = ""
            if info.state == "hidden":
                suffix = "（隐藏）"
            elif info.state == "veryHidden":
                suffix = "（深度隐藏）"
            ttk.Checkbutton(
                self.sheet_list.inner,
                text=f"{info.name} {suffix}",
                variable=var,
            ).grid(row=row, column=0, sticky="w", padx=16, pady=7)
        self.sheet_list.inner.columnconfigure(0, weight=1)

    def select_all(self) -> None:
        for var in self.sheet_vars.values():
            var.set(True)

    def clear_all(self) -> None:
        for var in self.sheet_vars.values():
            var.set(False)

    def selected_sheets(self) -> list[str]:
        return [name for name, var in self.sheet_vars.items() if var.get()]

    def set_running(self, value: bool) -> None:
        self.running = value
        state = "disabled" if value else "normal"
        for widget in (
            self.input_entry, self.output_entry, self.input_button, self.output_button,
            self.hidden_check, self.select_button, self.clear_button, self.start_button
        ):
            widget.configure(state=state)
        self.exit_button.configure(state="disabled" if value else "normal")

    def start_split(self) -> None:
        source = self.input_var.get().strip()
        output = self.output_var.get().strip()
        selected = self.selected_sheets()

        if not source:
            messagebox.showwarning("提示", "请选择源工作簿。")
            return
        if not output:
            messagebox.showwarning("提示", "请选择保存位置。")
            return
        if not selected:
            messagebox.showwarning("提示", "请至少选择一个工作表。")
            return

        self.progress["value"] = 0
        self.set_running(True)
        self.status_var.set("准备拆分...")
        threading.Thread(
            target=self.worker, args=(source, output, selected), daemon=True
        ).start()

    def worker(self, source: str, output: str, selected: list[str]) -> None:
        try:
            outputs = split_workbook(source, output, selected, self.thread_progress)
            self.after(0, self.success, outputs)
        except Exception as exc:
            detail = traceback.format_exc()
            self.after(0, self.failed, str(exc), detail)

    def thread_progress(self, completed: int, total: int, sheet: str, action: str) -> None:
        percent = int(completed / total * 100) if total else 0
        self.after(
            0,
            self.update_progress,
            percent,
            f"{action}：{sheet}（{completed}/{total}）",
        )

    def update_progress(self, percent: int, text: str) -> None:
        self.progress["value"] = percent
        self.status_var.set(text)

    def success(self, outputs: list[Path]) -> None:
        self.progress["value"] = 100
        self.set_running(False)
        self.status_var.set(f"拆分完成，共生成 {len(outputs)} 个文件。")
        if messagebox.askyesno(
            "拆分完成",
            f"已成功生成 {len(outputs)} 个文件。\n\n是否打开保存文件夹？",
        ):
            self.open_folder()

    def failed(self, error: str, detail: str) -> None:
        self.set_running(False)
        self.status_var.set("拆分失败。")
        messagebox.showerror(
            "拆分失败",
            f"{error}\n\n请确认 Excel 文件未被占用，且保存目录有写入权限。",
        )
        print(detail, file=sys.stderr)

    def open_folder(self) -> None:
        folder = Path(self.output_var.get().strip())
        folder.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(folder)  # type: ignore[attr-defined]
        except Exception:
            pass


def main() -> None:
    multiprocessing.freeze_support()
    App().mainloop()


if __name__ == "__main__":
    main()
